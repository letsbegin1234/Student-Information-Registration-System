from openpyxl import *
from tkinter import *

wb = load_workbook('Grades.xlsx')
sheet = wb.active

def excel():
	
	sheet.column_dimensions['A'].width = 30
	sheet.column_dimensions['B'].width = 10
	sheet.column_dimensions['C'].width = 10
	sheet.column_dimensions['D'].width = 20
	sheet.column_dimensions['E'].width = 20
	sheet.column_dimensions['F'].width = 40
	sheet.column_dimensions['G'].width = 50

	sheet.cell(row=1, column=1).value = "Name"
	sheet.cell(row=1, column=2).value = "Admission Number"
	sheet.cell(row=1, column=3).value = "Course"
	sheet.cell(row=1, column=4).value = "Branch"
	sheet.cell(row=1, column=5).value = "Contact Number"
	sheet.cell(row=1, column=6).value = "Email id"
	sheet.cell(row=1, column=7).value = "Address"

def focus1(event):
	course_field.focus_set()

def focus2(event):
	sem_field.focus_set()

def focus3(event):
	form_no_field.focus_set()

def focus4(event):
	contact_no_field.focus_set()

def focus5(event):
	email_id_field.focus_set()

def focus6(event):
	address_field.focus_set()

def clear():
	name_field.delete(0, END)
	course_field.delete(0, END)
	sem_field.delete(0, END)
	form_no_field.delete(0, END)
	contact_no_field.delete(0, END)
	email_id_field.delete(0, END)
	address_field.delete(0, END)

def insert():
	
	if (name_field.get() == "" and
		course_field.get() == "" and
		sem_field.get() == "" and
		form_no_field.get() == "" and
		contact_no_field.get() == "" and
		email_id_field.get() == "" and
		address_field.get() == ""):
			
		print("empty input")

	else:

		current_row = sheet.max_row
		current_column = sheet.max_column

		sheet.cell(row=current_row + 1, column=1).value = name_field.get()
		sheet.cell(row=current_row + 1, column=2).value = course_field.get()
		sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
		sheet.cell(row=current_row + 1, column=4).value = form_no_field.get()
		sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
		sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
		sheet.cell(row=current_row + 1, column=7).value = address_field.get()

		wb.save('Grades.xlsx')
		name_field.focus_set()
		print("Your data has been saved!!")

			
if __name__ == "__main__":
	
	root = Tk()
	root.configure(background='light green')
	root.title("Student Registration Form")
	root.geometry("600x400")
	excel()

	heading = Label(root, text="Student Details", font="bold", bg="light green")
	name = Label(root, text="Name",font="bold", bg="light green")
	course = Label(root, text="Admission Number",font="bold", bg="light green")
	sem = Label(root, text="Course", font="bold",bg="light green")
	form_no = Label(root, text="Branch", font="bold",bg="light green")
	contact_no = Label(root, text="Contact No.",font="bold", bg="light green")
	email_id = Label(root, text="Email id", font="bold",bg="light green")
	address = Label(root, text="Address", font="bold",bg="light green")

	heading.grid(row=0, column=1)
	name.grid(row=1, column=0)
	course.grid(row=2, column=0)
	sem.grid(row=3, column=0)
	form_no.grid(row=4, column=0)
	contact_no.grid(row=5, column=0)
	email_id.grid(row=6, column=0)
	address.grid(row=7, column=0)

	name_field = Entry(root)
	course_field = Entry(root)
	sem_field = Entry(root)
	form_no_field = Entry(root)
	contact_no_field = Entry(root)
	email_id_field = Entry(root)
	address_field = Entry(root)

	name_field.bind("<Return>", focus1)
	course_field.bind("<Return>", focus2)
	sem_field.bind("<Return>", focus3)
	form_no_field.bind("<Return>", focus4)
	contact_no_field.bind("<Return>", focus5)
	email_id_field.bind("<Return>", focus6)

	name_field.grid(row=1, column=1, ipadx="100")
	course_field.grid(row=2, column=1, ipadx="100")
	sem_field.grid(row=3, column=1, ipadx="100")
	form_no_field.grid(row=4, column=1, ipadx="100")
	contact_no_field.grid(row=5, column=1, ipadx="100")
	email_id_field.grid(row=6, column=1, ipadx="100")
	address_field.grid(row=7, column=1, ipadx="100")

	excel()

	submit = Button(root, text="Submit", fg="Black",bg="Red", command=insert)
	submit.grid(row=10, column=0)
	reset=Button(root, text="Reset", fg="Black", bg="Red", command=clear)
	reset.grid(row=10, column=3)

	root.mainloop()
